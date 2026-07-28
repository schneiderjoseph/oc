#!/usr/bin/env python3
"""
Audit complet base Optimum Control (schéma oc.*) → fichier Excel.

V2 : qualité fiches, doublons flous, amalgames proposés, plan d'action, factures POS enrichies.

Usage typique (accès direct SQL Server sur le poste La Réserve) :
  python oc_db_full_audit.py --server "NOM_SERVEUR\\INSTANCE" --database "NomBaseOC" --trusted

LocalDB (exercice) :
  python oc_db_full_audit.py --server "(localdb)\\mssqllocaldb" --database ocdata --trusted --config config.example.yaml
"""
from __future__ import annotations

import argparse
import re
import time
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal
from difflib import SequenceMatcher
from pathlib import Path

import pyodbc
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent
PROJECT_ROOT = BASE.parents[1]
VERSION = "2.0.0"
DEFAULT_OUT = PROJECT_ROOT / "output" / f"OC_Audit_Complet_{date.today():%Y%m%d}.xlsx"

POS_STATUS = {
    0: "Valid (liee)",
    1: "Unlinked",
    2: "Mismatched",
    3: "Pending",
    4: "Ignored",
}

DEFAULT_CONFIG = {
    "client_name": "",
    "thresholds": {"stock_value_alert": 15_000, "fuzzy_duplicate_min_score": 0.82},
    "exclude_name_patterns": [],
    "spirit_keywords": ["GIN", "WHISK", "RUM", "VODKA", "TEQUILA", "COGNAC", "HENNESSY"],
    "locations_bar": ["Bar", "bar"],
}

ODBC_DRIVERS = [
    "ODBC Driver 18 for SQL Server",
    "ODBC Driver 17 for SQL Server",
    "ODBC Driver 13 for SQL Server",
    "SQL Server",
]

# --- Styles Excel ---
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FILL_P1 = PatternFill("solid", fgColor="FFC7CE")
FILL_P2 = PatternFill("solid", fgColor="FCE4D6")
FILL_P3 = PatternFill("solid", fgColor="FFF2CC")
FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Audit OC → Excel")
    p.add_argument("--server", help="Serveur SQL (ex. (localdb)\\mssqllocaldb ou HOST\\SQLEXPRESS)")
    p.add_argument("--database", help="Nom de la base OC")
    p.add_argument("--user", help="Utilisateur SQL (si pas --trusted)")
    p.add_argument("--password", help="Mot de passe SQL")
    p.add_argument("--trusted", action="store_true", help="Authentification Windows")
    p.add_argument("--connection-string", help="Chaîne ODBC complète (prioritaire)")
    p.add_argument("--output", "-o", type=Path, default=DEFAULT_OUT, help="Fichier .xlsx de sortie")
    p.add_argument("--store-id", type=int, default=None, help="Filtrer sur un magasin (StoreId)")
    p.add_argument("--config", type=Path, default=None, help="Fichier YAML client (optionnel)")
    p.add_argument("--list-drivers", action="store_true", help="Afficher les drivers ODBC installés")
    return p.parse_args()


def load_config(path: Path | None) -> dict:
    cfg = {**DEFAULT_CONFIG, "thresholds": dict(DEFAULT_CONFIG["thresholds"])}
    if not path:
        return cfg
    try:
        import yaml
    except ImportError as exc:
        raise SystemExit("PyYAML requis pour --config : pip install pyyaml") from exc
    data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    if data.get("client_name"):
        cfg["client_name"] = data["client_name"]
    if data.get("thresholds"):
        cfg["thresholds"].update(data["thresholds"])
    for k in ("exclude_name_patterns", "spirit_keywords", "locations_bar"):
        if data.get(k):
            cfg[k] = data[k]
    if data.get("store_id") is not None:
        cfg["store_id"] = data["store_id"]
    return cfg


def build_connection_string(args: argparse.Namespace) -> str:
    if args.connection_string:
        return args.connection_string
    if not args.server or not args.database:
        raise SystemExit("Indiquez --connection-string OU --server + --database")
    driver = next((d for d in ODBC_DRIVERS if d in pyodbc.drivers()), None)
    if not driver:
        raise SystemExit(f"Aucun driver ODBC SQL Server trouvé. Installés : {pyodbc.drivers()}")
    parts = [
        f"Driver={{{driver}}}",
        f"Server={args.server}",
        f"Database={args.database}",
    ]
    if args.trusted or (not args.user and not args.password):
        parts.append("Trusted_Connection=yes")
    else:
        parts.append(f"UID={args.user}")
        parts.append(f"PWD={args.password}")
    return ";".join(parts) + ";"


def connect(cs: str) -> pyodbc.Connection:
    try:
        return pyodbc.connect(cs, timeout=30)
    except pyodbc.Error as exc:
        raise SystemExit(f"Connexion impossible : {exc}") from exc


def fetch(conn: pyodbc.Connection, sql: str, params=()) -> list[dict]:
    cur = conn.cursor()
    cur.execute(sql, params)
    cols = [d[0] for d in cur.description]
    rows = []
    for raw in cur.fetchall():
        row = {}
        for k, v in zip(cols, raw):
            if isinstance(v, Decimal):
                v = float(v)
            elif isinstance(v, datetime):
                v = v.replace(tzinfo=None)
            elif v is None:
                v = ""
            row[k] = v
        rows.append(row)
    return rows


def write_sheet(wb: Workbook, title: str, headers: list[str], rows: list[dict], *, max_width=48):
    """Crée ou remplace une feuille à partir d'en-têtes + lignes dict."""
    name = title[:31]
    if name in wb.sheetnames:
        ws = wb[name]
        wb.remove(ws)
    ws = wb.create_sheet(name)
    ws.append(headers)
    for c in ws[1]:
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    for ri in range(2, ws.max_row + 1):
        for ci in range(1, len(headers) + 1):
            cell = ws.cell(ri, ci)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for i, h in enumerate(headers, 1):
        width = min(max_width, max(len(str(h)), 10))
        for r in rows[:200]:
            width = max(width, min(len(str(r.get(h, ""))), max_width))
        ws.column_dimensions[get_column_letter(i)].width = width + 2
    ws.freeze_panes = "A2"
    return ws


def color_priority_sheet(ws, headers: list[str], prio_col: str = "Priorité"):
    if prio_col not in headers:
        return
    idx = headers.index(prio_col) + 1
    for ri in range(2, ws.max_row + 1):
        prio = ws.cell(ri, idx).value
        fill = None
        if prio == "P1":
            fill = FILL_P1
        elif prio == "P2":
            fill = FILL_P2
        elif prio == "P3":
            fill = FILL_P3
        if fill:
            for ci in range(1, len(headers) + 1):
                ws.cell(ri, ci).fill = fill


def norm_name(s: str) -> str:
    s = re.sub(r"\s+", " ", (s or "").strip().upper())
    return s


def conversion_text(cs: dict) -> str:
    parts = []
    if cs.get("CaseQty") and cs.get("CaseUom"):
        parts.append(f"{cs['CaseQty']} {cs['CaseUom']}")
    if cs.get("PakQty") and cs.get("PakUom"):
        parts.append(f"{cs['PakQty']} {cs['PakUom']}")
    if cs.get("PakConversionFactor"):
        parts.append(f"factor={cs['PakConversionFactor']}")
    return " / ".join(parts)


def classify_anomalies(item_rows: list[dict]) -> list[dict]:
    """Détection P1/P2/P3 — logique alignée sur l'audit La Réserve (bar.csv)."""
    by_name = defaultdict(list)
    for r in item_rows:
        by_name[norm_name(r.get("Descrip", ""))].append(r)

    issues: list[dict] = []
    seen_dup: set[str] = set()

    for r in item_rows:
        name = str(r.get("Descrip", ""))
        loc = str(r.get("Emplacement", r.get("PrimaryLocation", "")))
        u1 = str(r.get("PurchaseUom", "")).lower()
        u2 = str(r.get("CaseUom", "")).lower()
        u3 = str(r.get("RecipeUom", "")).lower()
        conv = str(r.get("Conversion", ""))
        val = float(r.get("StockValue") or 0)
        nkey = norm_name(name)

        def add(prio, typ, impact):
            issues.append({
                "Priorité": prio,
                "Type": typ,
                "ItemId": r.get("ItemId", ""),
                "Item": name,
                "Emplacement": loc,
                "UOM achat": r.get("PurchaseUom", ""),
                "UOM case": r.get("CaseUom", ""),
                "UOM recette": r.get("RecipeUom", ""),
                "Conversion": conv,
                "Valeur $": round(val, 2),
                "Impact métier": impact,
            })

        if re.search(r"^\.{3,}|^\?|^\.\.\.|^M$|^E\d+$", name) or "????" in name:
            add("P1", "Nom invalide / placeholder", "Renommer ou désactiver — fausse les rapports.")
        if name.upper() == "BEEF" and "bottle" in (u1, u3):
            add("P1", "Produit vs unité incohérent", "Bœuf en bottle — coût et usage idéal faux.")
        if "GANT" in name.upper() and u3 == "gal":
            add("P1", "Conversion absurde", "Gants en gallons — valeur stock aberrante.")
        if not conv.strip() and u1 in ("batch", "marmite"):
            add("P1", "Prep/batch sans conversion", "Batch inutilisable pour usage idéal.")
        if "mg / bottle" in conv.lower():
            add("P1", "mg au lieu de ml", "Erreur d'unité — coût portion ×1000.")
        if conv.strip().upper() in ("MUSCADOR", "LIQUOR") or (conv.strip() == "1.12" and "DIMPLE" in name.upper()):
            add("P1", "Conversion non numérique", "Champ conversion incohérent.")
        if val >= 15_000:
            add("P1", "Valeur inventaire aberrante", f"Stock valorisé {val:,.0f} $ — vérifier case size / conversion.")

        loc_l = loc.lower()
        if loc_l == "bar" or "bar" in loc_l.split(","):
            spirit_kw = ("GIN", "WHISK", "COGNAC", "HENNESSY", "RUM", "VODKA", "TEQUILA", "APEROL", "PINCH", "DSP")
            if any(k in name.upper() for k in spirit_kw):
                ok = u1 == "bottle" and u3 in ("oz (fl)", "ml", "oz")
                if not ok:
                    add("P2", "Spiritueux bar — mauvaise UOM", "Configurer bottle → oz (fl) ou ml (750).")

        if "cooler" in loc_l and u1 == "ea":
            if any(k in name.upper() for k in ("CHATEAU", "VIN", "WINE", "MERLOT", "MOET", "PROSECCO")):
                add("P2", "Vin en 'ea' au cooler", "Vins en bottle + ml pour coût verre.")

        if "*" in conv and "/" not in conv:
            add("P2", "Conversion format 1*XX", "Format non standard — risque erreur comptage.")
        if conv.strip() == "1*1" and "ml" in u3:
            add("P2", "Spiritueux 1*1 suspect", "Taille bouteille non définie (attendu 750 ml).")

        if nkey and nkey not in seen_dup and len(by_name[nkey]) > 1:
            locs = {str(x.get("Emplacement", x.get("PrimaryLocation", ""))) for x in by_name[nkey]}
            ids = {x.get("ItemId") for x in by_name[nkey]}
            if len(ids) > 1:
                configs = {
                    (x.get("PurchaseUom"), x.get("CaseUom"), x.get("RecipeUom"), x.get("Conversion"))
                    for x in by_name[nkey]
                }
                prio = "P2" if len(locs) >= 4 or len(configs) > 1 else "P3"
                total_val = sum(float(x.get("StockValue") or 0) for x in by_name[nkey])
                issues.append({
                    "Priorité": prio,
                    "Type": f"Doublon nom — {len(ids)} fiches",
                    "ItemId": ", ".join(str(i) for i in sorted(ids)),
                    "Item": name,
                    "Emplacement": ", ".join(sorted(l for l in locs if l)),
                    "UOM achat": r.get("PurchaseUom", ""),
                    "UOM case": r.get("CaseUom", ""),
                    "UOM recette": r.get("RecipeUom", ""),
                    "Conversion": f"{len(configs)} config(s)",
                    "Valeur $": round(total_val, 2),
                    "Impact métier": "Candidat amalgame — valider liste avant fusion.",
                })
                seen_dup.add(nkey)

    uniq: dict[tuple, dict] = {}
    for i in issues:
        k = (i["Priorité"], i["Type"], i["Item"], i["Emplacement"])
        if k not in uniq or i["Valeur $"] > uniq[k]["Valeur $"]:
            uniq[k] = i
    return sorted(uniq.values(), key=lambda x: (x["Priorité"], -x["Valeur $"], x["Item"]))


def fuzzy_duplicate_groups(item_rows: list[dict], min_score: float = 0.82) -> list[dict]:
    """Doublons probables par similarite de nom (difflib, sans dependance externe)."""
    entries = [(r["ItemId"], r.get("Descrip", ""), norm_name(r.get("Descrip", ""))) for r in item_rows]
    entries = [e for e in entries if e[2] and len(e[2]) >= 4]
    blocks: dict[str, list] = defaultdict(list)
    for e in entries:
        blocks[e[2][:4]].append(e)

    groups: list[dict] = []
    seen_pairs: set[tuple[int, int]] = set()
    for block in blocks.values():
        if len(block) < 2:
            continue
        for i in range(len(block)):
            for j in range(i + 1, len(block)):
                id_a, name_a, norm_a = block[i]
                id_b, name_b, norm_b = block[j]
                if norm_a == norm_b:
                    continue
                ratio = SequenceMatcher(None, norm_a, norm_b).ratio()
                if ratio < min_score:
                    continue
                pair = tuple(sorted((id_a, id_b)))
                if pair in seen_pairs:
                    continue
                seen_pairs.add(pair)
                keep_id = min(id_a, id_b)
                merge_ids = [x for x in (id_a, id_b) if x != keep_id]
                groups.append({
                    "Score": round(ratio, 3),
                    "Item garder": keep_id,
                    "Items fusionner": ", ".join(str(x) for x in merge_ids),
                    "Nom A": name_a,
                    "Nom B": name_b,
                    "Action": f"Valider amalgame -> garder ID {keep_id}",
                })
    return sorted(groups, key=lambda x: (-x["Score"], x["Nom A"]))


def build_quality_issues(item_rows: list[dict], product_rows: list[dict], prep_rows: list[dict]) -> list[dict]:
    issues = []
    product_ids = {r["ItemId"] for r in product_rows}
    prep_ids = {r["ItemId"] for r in prep_rows}
    products_by_id = {r["ItemId"]: r for r in product_rows}
    preps_by_id = {r["ItemId"]: r for r in prep_rows}

    for r in item_rows:
        iid = r.get("ItemId")
        name = r.get("Descrip", "")
        typ = r.get("Type", "")
        if typ == "I":
            if not r.get("Supplier"):
                issues.append({"Priorite": "P2", "Type": "Item sans fournisseur", "ItemId": iid, "Item": name,
                               "Detail": "Aucun CaseSize / fournisseur par defaut"})
            if not r.get("RecipeUom"):
                issues.append({"Priorite": "P2", "Type": "Item sans UOM recette", "ItemId": iid, "Item": name, "Detail": ""})
            if not r.get("PrimaryLocation") and not r.get("SecondaryLocations"):
                issues.append({"Priorite": "P3", "Type": "Item sans emplacement", "ItemId": iid, "Item": name, "Detail": ""})
        if r.get("IsDisabled") and typ in ("I", "P"):
            issues.append({"Priorite": "P3", "Type": "Fiche desactivee", "ItemId": iid, "Item": name, "Detail": typ})

    for r in product_rows:
        if not r.get("PosId"):
            issues.append({"Priorite": "P1", "Type": "Product sans POS ID", "ItemId": r["ItemId"],
                           "Item": r.get("Descrip", ""), "Detail": "Import POS -> Unlinked"})
        if int(r.get("IngredientCount") or 0) == 0:
            issues.append({"Priorite": "P1", "Type": "Product sans ingredients", "ItemId": r["ItemId"],
                           "Item": r.get("Descrip", ""), "Detail": ""})

    for r in prep_rows:
        if int(r.get("IngredientCount") or 0) == 0:
            issues.append({"Priorite": "P1", "Type": "Prep sans ingredients", "ItemId": r["ItemId"],
                           "Item": r.get("Descrip", ""), "Detail": ""})
        if not r.get("BatchUom"):
            issues.append({"Priorite": "P2", "Type": "Prep sans batch UOM", "ItemId": r["ItemId"],
                           "Item": r.get("Descrip", ""), "Detail": ""})

    return issues


def build_amalgamation_proposals(dup_rows: list[dict], fuzzy_rows: list[dict]) -> list[dict]:
    proposals = []
    for d in dup_rows:
        ids = [int(x.strip()) for x in str(d.get("ItemIds", "")).split(",") if x.strip().isdigit()]
        if not ids:
            continue
        keep = min(ids)
        merge = [x for x in ids if x != keep]
        proposals.append({
            "Source": "Nom exact",
            "Garder ItemId": keep,
            "Fusionner ItemIds": ", ".join(str(x) for x in merge),
            "Nom": d.get("NomNormalise", d.get("Variantes", "")),
            "Nb fiches": d.get("NbFiches", len(ids)),
            "Statut": "A valider par client",
        })
    for f in fuzzy_rows[:200]:
        proposals.append({
            "Source": f"Similarite {f['Score']}",
            "Garder ItemId": f["Item garder"],
            "Fusionner ItemIds": f["Items fusionner"],
            "Nom": f"{f['Nom A']} / {f['Nom B']}",
            "Nb fiches": 2,
            "Statut": "A valider par client",
        })
    return proposals


def build_action_plan(p1_count: int, p2_count: int, unlinked_count: int, dup_count: int) -> list[dict]:
    return [
        {"Semaine": "1", "Phase": "Audit + sauvegarde", "Actions": "Backup OC. Corriger P1 (" + str(p1_count) + "). Valider liste amalgames.",
         "Livrable": "Rapport P1 + liste fusions"},
        {"Semaine": "2", "Phase": "Bar / alcools", "Actions": "Unités bottle/oz/ml sur spiritueux. Corriger P2 bar (" + str(p2_count) + ").",
         "Livrable": "Bar prioritaire OK"},
        {"Semaine": "3", "Phase": "Amalgamations", "Actions": "Fusionner " + str(dup_count) + " groupes de doublons validés. 2e backup avant fusion.",
         "Livrable": "Doublons réduits"},
        {"Semaine": "4", "Phase": "POS + formation", "Actions": "Traiter ventes non liees (" + str(unlinked_count) + "). Daily Sales. Formation equipe.",
         "Livrable": "0 Unlinked test + guide"},
    ]


def label_sales_status(rows: list[dict]) -> list[dict]:
    out = []
    for r in rows:
        code = r.get("Status", "")
        try:
            label = POS_STATUS.get(int(code), f"Inconnu ({code})")
        except (TypeError, ValueError):
            label = str(code)
        out.append({**r, "Statut libelle": label})
    return out


SQL_STORE = """
SELECT StoreId, Identifier, Name, BusinessName, Address, City, Province, PostalCode,
       Country, Phone, ContactFirstName, ContactLastName, ContactEmail, Comments
FROM oc.Store
ORDER BY StoreId
"""

SQL_PREFERENCES = """
SELECT p.Name, p.Value, s.Name AS StoreName
FROM oc.Preference p
LEFT JOIN oc.Store s ON s.StoreId = p.Store
ORDER BY p.Store, p.Name
"""

SQL_SUPPLIERS = """
SELECT sup.SupplierId, sup.Name, sup.Active, sup.Address, sup.City, sup.Province,
       sup.PostalCode, sup.Country, sup.Phone,
       sup.ContactFirstName, sup.ContactLastName, sup.ContactEmail,
       sa.AccountNumber, sa.AccountReference
FROM oc.Supplier sup
LEFT JOIN oc.SupplierAccount sa ON sa.Supplier = sup.SupplierId
ORDER BY sup.Name
"""

SQL_SUPPLIER_ITEMS = """
SELECT sup.Name AS Fournisseur,
       i.ItemId, i.Type, i.Descrip,
       cs.OrderCode,
       cs.Descrip AS CaseSizeDescrip,
       pu.Uom AS PurchaseUom, cu.Uom AS CaseUom, pku.Uom AS PakUom,
       cs.CaseQty, cs.PakQty, cs.PakConversionFactor, cs.YieldFactor,
       csc.PurchasePrice, csc.UnitCost, csc.LastReceived
FROM oc.CaseSize cs
JOIN oc.Item i ON i.ItemId = cs.Item
JOIN oc.Supplier sup ON sup.SupplierId = cs.Supplier
LEFT JOIN oc.Uom pu ON pu.UomId = cs.PurchaseUom
LEFT JOIN oc.Uom cu ON cu.UomId = cs.CaseUom
LEFT JOIN oc.Uom pku ON pku.UomId = cs.PakUom
LEFT JOIN oc.CaseSizeCost csc ON csc.CaseSize = cs.CaseSizeId AND csc.IsDeleted = 0
ORDER BY sup.Name, i.Descrip
"""

SQL_LOCATIONS = """
SELECT l.LocationId, l.Name AS Location, s.StoreId, s.Name AS StoreName
FROM oc.Location l
JOIN oc.StoreLocation sl ON sl.Location = l.LocationId
JOIN oc.Store s ON s.StoreId = sl.Store
ORDER BY s.StoreId, l.Name
"""

SQL_ITEMS_MASTER = """
SELECT
    i.ItemId,
    i.Type,
    CASE i.Type WHEN 'I' THEN 'Item' WHEN 'P' THEN 'Prep' WHEN 'M' THEN 'Product' ELSE i.Type END AS TypeLabel,
    i.Descrip,
    g.Descrip AS InventoryGroup,
    cat.Name AS SalesCategory,
    ru.Uom AS RecipeUom,
    pu.Uom AS PurchaseUom,
    cu.Uom AS CaseUom,
    pku.Uom AS PakUom,
    cs.CaseQty,
    cs.PakQty,
    cs.PakConversionFactor,
    cs.OrderCode,
    sup.Name AS Supplier,
    csc.PurchasePrice,
    csc.UnitCost,
    q.QtyOnHand,
    COALESCE(iuc.TrackingCost, csc.UnitCost, 0) AS UnitCostUsed,
    q.QtyOnHand * COALESCE(iuc.TrackingCost, csc.UnitCost, 0) AS StockValue,
    ploc.Name AS PrimaryLocation,
    STUFF((
        SELECT ', ' + l2.Name
        FROM oc.ItemLocation il2
        JOIN oc.Location l2 ON l2.LocationId = il2.Location
        WHERE il2.Item = i.ItemId
          AND (kid.PrimaryLocation IS NULL OR il2.Location <> kid.PrimaryLocation)
        ORDER BY il2.SortIdx
        FOR XML PATH(''), TYPE
    ).value('.', 'NVARCHAR(MAX)'), 1, 2, '') AS SecondaryLocations,
    id.TrackInventory,
    id.ActualizeUsage,
    cs.YieldFactor,
    CASE WHEN de.EntityId IS NOT NULL THEN 1 ELSE 0 END AS IsDisabled
FROM oc.Item i
LEFT JOIN oc.[Group] g ON g.GroupId = i.ItemGroup
LEFT JOIN oc.Category cat ON cat.CategoryId = g.Category
LEFT JOIN oc.Uom ru ON ru.UomId = i.RecipeUom
LEFT JOIN oc.ItemDetail id ON id.Item = i.ItemId
LEFT JOIN oc.KeyItemDetail kid ON kid.Item = i.ItemId
LEFT JOIN oc.Location ploc ON ploc.LocationId = kid.PrimaryLocation
LEFT JOIN oc.ItemQtyOnHand q ON q.Item = i.ItemId
LEFT JOIN oc.ItemUnitCost iuc ON iuc.Item = i.ItemId AND iuc.Active = 1
OUTER APPLY (
    SELECT TOP 1 cs.*
    FROM oc.CaseSize cs
    WHERE cs.Item = i.ItemId
    ORDER BY CASE WHEN cs.CaseSizeId = id.DefaultCaseSize THEN 0
                  WHEN cs.CaseSizeId = id.CurrentCaseSize THEN 1 ELSE 2 END,
             cs.CaseSizeId
) cs
LEFT JOIN oc.CaseSizeCost csc ON csc.CaseSize = cs.CaseSizeId AND csc.IsDeleted = 0
LEFT JOIN oc.Supplier sup ON sup.SupplierId = cs.Supplier
LEFT JOIN oc.Uom pu ON pu.UomId = cs.PurchaseUom
LEFT JOIN oc.Uom cu ON cu.UomId = cs.CaseUom
LEFT JOIN oc.Uom pku ON pku.UomId = cs.PakUom
LEFT JOIN oc.DisabledEntity de ON de.EntityId = i.ItemId AND de.EntityType = 1
ORDER BY i.Type, i.Descrip
"""

SQL_CASE_SIZES = """
SELECT i.ItemId, i.Descrip, i.Type, cs.CaseSizeId, cs.Descrip AS CaseDescrip,
       sup.Name AS Supplier, cs.OrderCode,
       pu.Uom AS PurchaseUom, cu.Uom AS CaseUom, pku.Uom AS PakUom,
       cs.CaseQty, cs.PakQty, cs.PakConversionFactor, cs.YieldFactor,
       csc.PurchasePrice, csc.CaseCost, csc.UnitCost, csc.LastReceived
FROM oc.CaseSize cs
JOIN oc.Item i ON i.ItemId = cs.Item
LEFT JOIN oc.Supplier sup ON sup.SupplierId = cs.Supplier
LEFT JOIN oc.Uom pu ON pu.UomId = cs.PurchaseUom
LEFT JOIN oc.Uom cu ON cu.UomId = cs.CaseUom
LEFT JOIN oc.Uom pku ON pku.UomId = cs.PakUom
LEFT JOIN oc.CaseSizeCost csc ON csc.CaseSize = cs.CaseSizeId AND csc.IsDeleted = 0
ORDER BY i.Descrip, cs.CaseSizeId
"""

SQL_CONVERSIONS = """
SELECT i.ItemId, i.Descrip, fu.Uom AS FromUom, c.FromQty, tu.Uom AS ToUom, c.ToQty, c.Multiplier
FROM oc.Conversion c
JOIN oc.Item i ON i.ItemId = c.Item
LEFT JOIN oc.Uom fu ON fu.UomId = c.FromUom
LEFT JOIN oc.Uom tu ON tu.UomId = c.ToUom
ORDER BY i.Descrip
"""

SQL_PREPS = """
SELECT i.ItemId, i.Descrip, ru.Uom AS RecipeUom, bu.Uom AS BatchUom,
       py.CountQty, cu.Uom AS CountUom, py.RecipeQty,
       pd.ShowOnPrepSheets, ps.Description AS PrepStation, pd.ShelfLifeMinutes,
       id.ActualizeUsage,
       (SELECT COUNT(*) FROM oc.Ingredient ing WHERE ing.Recipe = i.ItemId) AS IngredientCount
FROM oc.Item i
JOIN oc.PrepYield py ON py.Item = i.ItemId
JOIN oc.Uom ru ON ru.UomId = i.RecipeUom
JOIN oc.Uom bu ON bu.UomId = py.BatchUom
JOIN oc.Uom cu ON cu.UomId = py.CountUom
LEFT JOIN oc.PrepDetail pd ON pd.Item = i.ItemId
LEFT JOIN oc.PrepStation ps ON ps.PrepStationId = pd.PrepStation
LEFT JOIN oc.ItemDetail id ON id.Item = i.ItemId
ORDER BY i.Descrip
"""

SQL_INGREDIENTS = """
SELECT parent.Descrip AS Recipe, parent.Type AS RecipeType, parent.ItemId AS RecipeId,
       ing.Idx, child.Descrip AS Ingredient, child.Type AS IngredientType,
       child.ItemId AS IngredientId, ing.Qty, u.Uom
FROM oc.Ingredient ing
JOIN oc.Item parent ON parent.ItemId = ing.Recipe
JOIN oc.Item child ON child.ItemId = ing.Item
JOIN oc.Uom u ON u.UomId = ing.Uom
ORDER BY parent.Descrip, ing.Idx
"""

SQL_PRODUCTS = """
SELECT i.ItemId, i.Descrip, pp.Plu AS PosId, pp.PluDescrip, pp.Price, pp.Active,
       g.Descrip AS SalesGroup, cat.Name AS SalesCategory,
       (SELECT COUNT(*) FROM oc.Ingredient ing WHERE ing.Recipe = i.ItemId) AS IngredientCount
FROM oc.Item i
JOIN oc.ProductPrice pp ON pp.Item = i.ItemId
LEFT JOIN oc.[Group] g ON g.GroupId = i.ItemGroup
LEFT JOIN oc.Category cat ON cat.CategoryId = g.Category
WHERE i.Type = 'M'
ORDER BY pp.Plu, i.Descrip
"""

SQL_DUPLICATE_NAMES = """
SELECT UPPER(LTRIM(RTRIM(i.Descrip))) AS NomNormalise,
       COUNT(DISTINCT i.ItemId) AS NbFiches,
       STRING_AGG(CAST(i.ItemId AS VARCHAR(12)), ', ') WITHIN GROUP (ORDER BY i.ItemId) AS ItemIds,
       STRING_AGG(i.Descrip, ' | ') WITHIN GROUP (ORDER BY i.ItemId) AS Variantes
FROM oc.Item i
WHERE i.Descrip IS NOT NULL AND LTRIM(RTRIM(i.Descrip)) <> ''
GROUP BY UPPER(LTRIM(RTRIM(i.Descrip)))
HAVING COUNT(DISTINCT i.ItemId) > 1
ORDER BY COUNT(DISTINCT i.ItemId) DESC, UPPER(LTRIM(RTRIM(i.Descrip)))
"""

SQL_INVOICES_SUMMARY = """
SELECT sup.Name AS Fournisseur,
       COUNT(*) AS NbFactures,
       MIN(inv.InvoiceDate) AS PremiereFacture,
       MAX(inv.InvoiceDate) AS DerniereFacture,
       SUM(inv.Total) AS TotalTTC
FROM oc.Invoice inv
LEFT JOIN oc.Supplier sup ON sup.SupplierId = inv.Supplier
GROUP BY sup.Name
ORDER BY SUM(inv.Total) DESC
"""

SQL_INVOICES_RECENT = """
SELECT TOP 500 inv.InvoiceNumber, inv.InvoiceDate, sup.Name AS Fournisseur,
       inv.Total, inv.ItemTotal, inv.AdjustmentTotal
FROM oc.Invoice inv
LEFT JOIN oc.Supplier sup ON sup.SupplierId = inv.Supplier
ORDER BY inv.InvoiceDate DESC, inv.InvoiceId DESC
"""

SQL_SALES_STATUS = """
SELECT si.Status,
       COUNT(*) AS NbLignes,
       SUM(si.QtySold) AS QtyTotale,
       SUM(si.GrossSales) AS VentesBrutes
FROM oc.SalesItem si
GROUP BY si.Status
ORDER BY si.Status
"""

SQL_SALES_UNLINKED = """
SELECT TOP 500 ss.SalesDate, si.PluNumber, si.Descrip, si.SalesGroup,
       si.QtySold, si.GrossSales, si.Status, si.Department
FROM oc.SalesItem si
JOIN oc.SalesSource ss ON ss.SalesSourceId = si.SalesSource
WHERE si.Status NOT IN (0)
ORDER BY ss.SalesDate DESC, si.PluNumber
"""

SQL_INVENTORIES = """
SELECT inv.InventoryId, inv.OpenDate, inv.CloseDate, inv.Finalized,
       s.Name AS StoreName
FROM oc.Inventory inv
LEFT JOIN oc.Store s ON s.StoreId = inv.Store
ORDER BY inv.InventoryId DESC
"""

SQL_TABLE_COUNTS = """
SELECT t.name AS TableName, SUM(p.rows) AS TotalRows
FROM sys.tables t
JOIN sys.partitions p ON p.object_id = t.object_id AND p.index_id IN (0, 1)
WHERE SCHEMA_NAME(t.schema_id) = 'oc'
GROUP BY t.name
ORDER BY SUM(p.rows) DESC, t.name
"""

SQL_INVOICE_IMBALANCE = """
SELECT TOP 200 inv.InvoiceNumber, inv.InvoiceDate, sup.Name AS Fournisseur,
       inv.Total, inv.ItemTotal, inv.AdjustmentTotal, inv.ExpenseTotal,
       inv.Total - inv.ItemTotal - inv.AdjustmentTotal - inv.ExpenseTotal AS Balance
FROM oc.Invoice inv
LEFT JOIN oc.Supplier sup ON sup.SupplierId = inv.Supplier
WHERE ABS(inv.Total - inv.ItemTotal - inv.AdjustmentTotal - inv.ExpenseTotal) > 0.02
ORDER BY inv.InvoiceDate DESC
"""

SQL_INVOICE_ZERO_COST = """
SELECT TOP 200 inv.InvoiceNumber, inv.InvoiceDate, i.Descrip, ii.Qty, ii.UnitCost, ii.LineTotal
FROM oc.InvoiceItem ii
JOIN oc.Invoice inv ON inv.InvoiceId = ii.Invoice
JOIN oc.Item i ON i.ItemId = ii.Item
WHERE ii.UnitCost = 0 OR ii.LineTotal = 0
ORDER BY inv.InvoiceDate DESC
"""

SQL_ORDER_CODE_DUPES = """
SELECT sup.Name AS Fournisseur, cs.OrderCode,
       COUNT(DISTINCT cs.Item) AS NbItems,
       STRING_AGG(CAST(i.ItemId AS VARCHAR(12)), ', ') WITHIN GROUP (ORDER BY i.ItemId) AS ItemIds,
       STRING_AGG(i.Descrip, ' | ') WITHIN GROUP (ORDER BY i.ItemId) AS Items
FROM oc.CaseSize cs
JOIN oc.Supplier sup ON sup.SupplierId = cs.Supplier
JOIN oc.Item i ON i.ItemId = cs.Item
WHERE cs.OrderCode IS NOT NULL AND LTRIM(RTRIM(cs.OrderCode)) <> ''
GROUP BY sup.Name, cs.OrderCode
HAVING COUNT(DISTINCT cs.Item) > 1
ORDER BY COUNT(DISTINCT cs.Item) DESC
"""

SQL_SALES_DAILY = """
SELECT CAST(ss.SalesDate AS date) AS Jour,
       SUM(CASE WHEN si.Status <> 0 THEN 1 ELSE 0 END) AS NonLiees,
       COUNT(*) AS TotalLignes
FROM oc.SalesItem si
JOIN oc.SalesSource ss ON ss.SalesSourceId = si.SalesSource
GROUP BY CAST(ss.SalesDate AS date)
ORDER BY CAST(ss.SalesDate AS date) DESC
"""


# ---------------------------------------------------------------------------
# Requêtes SQL (suite V1)
# ---------------------------------------------------------------------------


def enrich_items(rows: list[dict]) -> list[dict]:
    for r in rows:
        parts = []
        if r.get("CaseQty") and r.get("CaseUom"):
            parts.append(f"{r['CaseQty']} {r['CaseUom']}")
        if r.get("PakQty") and r.get("PakUom"):
            parts.append(f"{r['PakQty']} {r['PakUom']}")
        if r.get("PakConversionFactor"):
            parts.append(f"factor={r['PakConversionFactor']}")
        r["Conversion"] = " / ".join(parts)
        prim = str(r.get("PrimaryLocation", "") or "")
        sec = str(r.get("SecondaryLocations", "") or "")
        r["Emplacement"] = ", ".join(x for x in (prim, sec) if x)
    return rows


def build_resume(store_rows, item_rows, anomaly_rows, supplier_rows, dup_rows, sales_status,
                 quality_rows, fuzzy_rows, meta: dict) -> list[dict]:
    types = Counter(r.get("TypeLabel", r.get("Type", "")) for r in item_rows)
    p_counts = Counter(a["Priorité"] for a in anomaly_rows)
    q_p1 = sum(1 for q in quality_rows if q.get("Priorite") == "P1")
    labeled = label_sales_status(sales_status)
    return [
        {"Indicateur": "Version script", "Valeur": VERSION},
        {"Indicateur": "Client", "Valeur": meta.get("client_name") or "—"},
        {"Indicateur": "Date audit", "Valeur": date.today().isoformat()},
        {"Indicateur": "Duree (sec)", "Valeur": meta.get("duration_sec", "")},
        {"Indicateur": "Etablissement(s)", "Valeur": ", ".join(r.get("Name", "") for r in store_rows)},
        {"Indicateur": "Fournisseurs actifs", "Valeur": sum(1 for r in supplier_rows if r.get("Active"))},
        {"Indicateur": "Items (I)", "Valeur": types.get("Item", 0)},
        {"Indicateur": "Preps (P)", "Valeur": types.get("Prep", 0)},
        {"Indicateur": "Products (M)", "Valeur": types.get("Product", 0)},
        {"Indicateur": "Total fiches Item", "Valeur": len(item_rows)},
        {"Indicateur": "Doublons noms exacts", "Valeur": len(dup_rows)},
        {"Indicateur": "Doublons flous", "Valeur": len(fuzzy_rows)},
        {"Indicateur": "Anomalies P1", "Valeur": p_counts.get("P1", 0)},
        {"Indicateur": "Anomalies P2", "Valeur": p_counts.get("P2", 0)},
        {"Indicateur": "Anomalies P3", "Valeur": p_counts.get("P3", 0)},
        {"Indicateur": "Qualite P1", "Valeur": q_p1},
        {
            "Indicateur": "Ventes POS (statut)",
            "Valeur": ", ".join(
                f"{r.get('Statut libelle')}={r.get('NbLignes')}" for r in labeled
            ) if labeled else "—",
        },
    ]


def build_readme() -> list[dict]:
    return [
        {"Section": "Résumé", "Description": "Vue d'ensemble chiffrée"},
        {"Section": "Établissement", "Description": "Infos magasin / La Réserve (oc.Store)"},
        {"Section": "Préférences", "Description": "Paramètres système OC"},
        {"Section": "Fournisseurs", "Description": "Liste fournisseurs + comptes"},
        {"Section": "Fournisseur-Items", "Description": "Qui fournit quoi (CaseSize)"},
        {"Section": "Emplacements", "Description": "Storage locations"},
        {"Section": "Items complet", "Description": "Toutes les fiches avec UOM, stock, emplacement"},
        {"Section": "Case Sizes", "Description": "Tous les conditionnements / prix"},
        {"Section": "Conversions", "Description": "Conversions unités par item"},
        {"Section": "Preps", "Description": "Préparations batch"},
        {"Section": "Ingrédients", "Description": "Recettes (items/preps/products)"},
        {"Section": "Products", "Description": "Menu + POS ID"},
        {"Section": "Doublons noms", "Description": "Candidats amalgame (même nom, IDs différents)"},
        {"Section": "Anomalies", "Description": "P1 critique / P2 important / P3 doublons"},
        {"Section": "Factures résumé", "Description": "Achats par fournisseur"},
        {"Section": "Factures récentes", "Description": "500 dernières factures"},
        {"Section": "Ventes statut", "Description": "POS — comptage par Status"},
        {"Section": "Ventes à traiter", "Description": "Lignes non liées (Status ≠ 0)"},
        {"Section": "Inventaires", "Description": "Cycles d'inventaire"},
        {"Section": "Tables (volumes)", "Description": "Nombre de lignes par table oc.*"},
        {"Section": "Plan action", "Description": "Semaines 1-4 auto-generees"},
        {"Section": "Amalgames proposes", "Description": "Fusions a valider par le client"},
        {"Section": "Qualite fiches", "Description": "Items/preps/products incomplets"},
        {"Section": "Doublons flous", "Description": "Noms similaires (score >= seuil config)"},
        {"Section": "Order codes dupes", "Description": "Meme code fournisseur sur 2+ items"},
        {"Section": "Factures desequilibrees", "Description": "Total != lignes + ajustements"},
        {"Section": "Factures cout zero", "Description": "Lignes a 0 $"},
        {"Section": "Ventes par jour", "Description": "Unlinked par jour"},
        {"Section": "Metadonnees", "Description": "Version, duree, serveur"},
    ]


def run_audit(conn: pyodbc.Connection, out_path: Path, cfg: dict | None = None) -> None:
    cfg = cfg or DEFAULT_CONFIG
    t0 = time.time()
    print("Lecture base OC (V2)…")
    store_rows = fetch(conn, SQL_STORE)
    pref_rows = fetch(conn, SQL_PREFERENCES)
    supplier_rows = fetch(conn, SQL_SUPPLIERS)
    supplier_items = fetch(conn, SQL_SUPPLIER_ITEMS)
    location_rows = fetch(conn, SQL_LOCATIONS)
    item_rows = enrich_items(fetch(conn, SQL_ITEMS_MASTER))
    case_rows = fetch(conn, SQL_CASE_SIZES)
    conv_rows = fetch(conn, SQL_CONVERSIONS)
    prep_rows = fetch(conn, SQL_PREPS)
    ing_rows = fetch(conn, SQL_INGREDIENTS)
    product_rows = fetch(conn, SQL_PRODUCTS)

    try:
        dup_rows = fetch(conn, SQL_DUPLICATE_NAMES)
    except pyodbc.Error:
        dup_rows = []
        by_n = defaultdict(list)
        for r in item_rows:
            by_n[norm_name(r.get("Descrip", ""))].append(r)
        for n, grp in by_n.items():
            if n and len({x["ItemId"] for x in grp}) > 1:
                dup_rows.append({
                    "NomNormalise": n,
                    "NbFiches": len({x["ItemId"] for x in grp}),
                    "ItemIds": ", ".join(str(x["ItemId"]) for x in sorted(grp, key=lambda z: z["ItemId"])),
                    "Variantes": " | ".join(x["Descrip"] for x in grp[:5]),
                })

    min_fuzzy = float(cfg["thresholds"].get("fuzzy_duplicate_min_score", 0.82))
    fuzzy_rows = fuzzy_duplicate_groups(item_rows, min_fuzzy)
    quality_rows = build_quality_issues(item_rows, product_rows, prep_rows)
    amalgam_rows = build_amalgamation_proposals(dup_rows, fuzzy_rows)
    anomaly_rows = classify_anomalies(item_rows)
    inv_sum = fetch(conn, SQL_INVOICES_SUMMARY)
    inv_recent = fetch(conn, SQL_INVOICES_RECENT)
    sales_status = fetch(conn, SQL_SALES_STATUS)
    sales_status_l = label_sales_status(sales_status)
    sales_open = fetch(conn, SQL_SALES_UNLINKED)
    sales_daily = fetch(conn, SQL_SALES_DAILY)
    inventories = fetch(conn, SQL_INVENTORIES)
    table_counts = fetch(conn, SQL_TABLE_COUNTS)

    try:
        inv_imbalance = fetch(conn, SQL_INVOICE_IMBALANCE)
    except pyodbc.Error:
        inv_imbalance = []
    try:
        inv_zero = fetch(conn, SQL_INVOICE_ZERO_COST)
    except pyodbc.Error:
        inv_zero = []
    try:
        order_dupes = fetch(conn, SQL_ORDER_CODE_DUPES)
    except pyodbc.Error:
        order_dupes = []

    p1 = sum(1 for a in anomaly_rows if a["Priorité"] == "P1") + sum(1 for q in quality_rows if q.get("Priorite") == "P1")
    p2 = sum(1 for a in anomaly_rows if a["Priorité"] == "P2") + sum(1 for q in quality_rows if q.get("Priorite") == "P2")
    unlinked = sum(int(r.get("NbLignes") or 0) for r in sales_status_l if r.get("Status") not in (0, "0"))
    action_plan = build_action_plan(p1, p2, unlinked, len(dup_rows) + len(fuzzy_rows))

    duration = round(time.time() - t0, 1)
    meta = {
        "client_name": cfg.get("client_name", ""),
        "duration_sec": duration,
        "version": VERSION,
        "generated": datetime.now().isoformat(timespec="seconds"),
    }
    meta_rows = [{"Cle": k, "Valeur": v} for k, v in meta.items()]

    wb = Workbook()
    wb.remove(wb.active)

    sheets = [
        ("README", ["Section", "Description"], build_readme()),
        ("Resume", ["Indicateur", "Valeur"], build_resume(
            store_rows, item_rows, anomaly_rows, supplier_rows, dup_rows, sales_status,
            quality_rows, fuzzy_rows, meta)),
        ("Plan action", ["Semaine", "Phase", "Actions", "Livrable"], action_plan),
        ("Amalgames proposes", list(amalgam_rows[0].keys()) if amalgam_rows else ["Garder ItemId"], amalgam_rows),
        ("Metadonnees", ["Cle", "Valeur"], meta_rows),
        ("Qualite fiches", list(quality_rows[0].keys()) if quality_rows else ["Priorite", "Type", "Item"], quality_rows),
        ("Doublons flous", list(fuzzy_rows[0].keys()) if fuzzy_rows else ["Score", "Nom A"], fuzzy_rows),
        ("Etablissement", list(store_rows[0].keys()) if store_rows else ["Info"], store_rows),
        ("Preferences", list(pref_rows[0].keys()) if pref_rows else ["Name", "Value"], pref_rows),
        ("Fournisseurs", list(supplier_rows[0].keys()) if supplier_rows else ["SupplierId"], supplier_rows),
        ("Fournisseur-Items", list(supplier_items[0].keys()) if supplier_items else ["Fournisseur"], supplier_items),
        ("Emplacements", list(location_rows[0].keys()) if location_rows else ["Location"], location_rows),
        ("Items complet", list(item_rows[0].keys()) if item_rows else ["ItemId"], item_rows),
        ("Case Sizes", list(case_rows[0].keys()) if case_rows else ["CaseSizeId"], case_rows),
        ("Conversions", list(conv_rows[0].keys()) if conv_rows else ["ItemId"], conv_rows),
        ("Preps", list(prep_rows[0].keys()) if prep_rows else ["ItemId"], prep_rows),
        ("Ingredients", list(ing_rows[0].keys()) if ing_rows else ["Recipe"], ing_rows),
        ("Products", list(product_rows[0].keys()) if product_rows else ["ItemId"], product_rows),
        ("Doublons noms", list(dup_rows[0].keys()) if dup_rows else ["NomNormalise"], dup_rows),
        ("Order codes dupes", list(order_dupes[0].keys()) if order_dupes else ["Fournisseur"], order_dupes),
        ("Anomalies", list(anomaly_rows[0].keys()) if anomaly_rows else [
            "Priorité", "Type", "ItemId", "Item", "Emplacement", "Impact métier"
        ], anomaly_rows),
        ("Factures resume", list(inv_sum[0].keys()) if inv_sum else ["Fournisseur"], inv_sum),
        ("Factures recentes", list(inv_recent[0].keys()) if inv_recent else ["InvoiceNumber"], inv_recent),
        ("Factures desequilibrees", list(inv_imbalance[0].keys()) if inv_imbalance else ["InvoiceNumber"], inv_imbalance),
        ("Factures cout zero", list(inv_zero[0].keys()) if inv_zero else ["InvoiceNumber"], inv_zero),
        ("Ventes statut", list(sales_status_l[0].keys()) if sales_status_l else ["Status"], sales_status_l),
        ("Ventes a traiter", list(sales_open[0].keys()) if sales_open else ["PluNumber"], sales_open),
        ("Ventes par jour", list(sales_daily[0].keys()) if sales_daily else ["Jour"], sales_daily),
        ("Inventaires", list(inventories[0].keys()) if inventories else ["InventoryId"], inventories),
        ("Tables volumes", list(table_counts[0].keys()) if table_counts else ["TableName"], table_counts),
    ]

    for title, headers, rows in sheets:
        print(f"  -> {title} ({len(rows)} lignes)")
        ws = write_sheet(wb, title, headers, rows)
        if title in ("Anomalies", "Qualite fiches"):
            color_priority_sheet(ws, headers, prio_col="Priorité" if title == "Anomalies" else "Priorite")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"\nFichier cree : {out_path} ({duration}s)")


def main() -> None:
    args = parse_args()
    if args.list_drivers:
        print("Drivers ODBC installés :")
        for d in pyodbc.drivers():
            print(f"  - {d}")
        return
    cs = build_connection_string(args)
    cfg = load_config(args.config)
    if args.store_id is not None:
        cfg["store_id"] = args.store_id
    print(f"Connexion : {cs.split('PWD=')[0]}…")
    conn = connect(cs)
    try:
        run_audit(conn, args.output, cfg)
    finally:
        conn.close()


if __name__ == "__main__":
    main()
